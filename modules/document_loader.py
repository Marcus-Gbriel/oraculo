"""
Módulo de Carregamento de Documentos
Responsável por ler e extrair texto de documentos PDF, DOCX e Excel
"""

import os
from typing import List, Dict, Optional
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class DocumentLoader:
    """
    Classe para carregar e extrair texto de diversos formatos de documentos
    """
    
    def __init__(self, documents_path: str):
        """
        Inicializa o carregador de documentos
        
        Args:
            documents_path: Caminho para a pasta com os documentos
        """
        self.documents_path = Path(documents_path)
        self.supported_formats = ['.pdf', '.docx', '.xlsx', '.xls']
        
    def load_pdf(self, file_path: Path) -> str:
        """
        Carrega e extrai texto de um arquivo PDF
        
        Args:
            file_path: Caminho do arquivo PDF
            
        Returns:
            Texto extraído do PDF
        """
        try:
            import PyPDF2
            
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            
            logger.info(f"[PDF] Carregado com sucesso: {file_path.name}")
            return text
        except Exception as e:
            logger.error(f"[ERRO] Erro ao carregar PDF {file_path.name}: {str(e)}")
            return ""
    
    def load_docx(self, file_path: Path) -> str:
        """
        Carrega e extrai texto de um arquivo DOCX
        
        Args:
            file_path: Caminho do arquivo DOCX
            
        Returns:
            Texto extraído do DOCX
        """
        try:
            from docx import Document
            
            doc = Document(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
            logger.info(f"[DOCX] Carregado com sucesso: {file_path.name}")
            return text
        except Exception as e:
            logger.error(f"[ERRO] Erro ao carregar DOCX {file_path.name}: {str(e)}")
            return ""
    
    def load_excel(self, file_path: Path) -> str:
        """
        Carrega e extrai texto de um arquivo Excel
        
        Args:
            file_path: Caminho do arquivo Excel
            
        Returns:
            Texto extraído do Excel
        """
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n--- Planilha: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            
            logger.info(f"[EXCEL] Carregado com sucesso: {file_path.name}")
            return text
        except Exception as e:
            logger.error(f"[ERRO] Erro ao carregar Excel {file_path.name}: {str(e)}")
            return ""
    
    def load_all_documents(self) -> List[Dict[str, str]]:
        """
        Carrega todos os documentos suportados da pasta
        
        Returns:
            Lista de dicionários contendo nome do arquivo e conteúdo
        """
        documents = []
        
        if not self.documents_path.exists():
            logger.error(f"[ERRO] Pasta nao encontrada: {self.documents_path}")
            return documents
        
        for file_path in self.documents_path.iterdir():
            if file_path.is_file() and file_path.suffix.lower() in self.supported_formats:
                content = ""
                
                if file_path.suffix.lower() == '.pdf':
                    content = self.load_pdf(file_path)
                elif file_path.suffix.lower() == '.docx':
                    content = self.load_docx(file_path)
                elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                    content = self.load_excel(file_path)
                
                if content:
                    documents.append({
                        'filename': file_path.name,
                        'content': content,
                        'path': str(file_path)
                    })
                    logger.info(f"[DOCUMENTO] Processado: {file_path.name} ({len(content)} caracteres)")
        
        logger.info(f"[DOCUMENTO] Total carregado: {len(documents)} documentos")
        return documents
